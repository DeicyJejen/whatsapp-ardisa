import subprocess, os, re
from collections import Counter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

BASE='/home/ubuntu/whatsapp-ardisa/'; OUT=BASE+'reportes/'; TMP='/tmp/rep_logos/'
os.makedirs(OUT, exist_ok=True); os.makedirs(TMP, exist_ok=True)
NAVY='1E2A4A'; TEAL='0F9D8E'; AMBER='F5B301'; LIGHT='F4F6F8'; GREY='5A6472'
EMOJI=re.compile('[\U0001F000-\U0001FAFF\U00002600-\U000027BF\U0001F1E6-\U0001F1FF\U00002B00-\U00002BFF\U00002190-\U000021FF\U0000FE0F]')
def clean(s): return re.sub(r'\s+',' ',EMOJI.sub('', str(s or ''))).strip()
LOGO_DIRS=[BASE+'oficina/', BASE]   # fix 2026-08-03: los logos se movieron a oficina/ el 22-jul
def ruta_logo(nombre):
    for d in LOGO_DIRS:
        if os.path.exists(d+nombre): return d+nombre
    print('!! LOGO NO ENCONTRADO: '+nombre); return LOGO_DIRS[-1]+nombre
def logo_png(marca):
    src=ruta_logo('logofirmagrupoardisavertical_org.png' if marca=='Ardisa' else 'logofirmacarpincentrovertical2.png')
    im=PILImage.open(src).convert('RGBA'); bb=im.getbbox()
    if bb: im=im.crop(bb)
    bg=PILImage.new('RGB', im.size, 'white'); bg.paste(im, mask=im.split()[3])
    o=TMP+'x_'+marca+'.png'; bg.save(o,'PNG'); return o, bg.size
def q(sql):
    out=subprocess.check_output(['sudo','mysql','--default-character-set=utf8mb4','bot_ardisa','-N','-B','-e',sql], encoding='utf-8', errors='replace')
    return [ln.split('\t') for ln in out.strip().split('\n') if ln.strip()]

zones=q("SELECT DISTINCT marca,ciudad FROM leads ORDER BY marca,ciudad")
COLS=['Fecha','Cliente','Telefono','Tipo de cliente','Solicitud','Detalle','Asesor que atendio']
WRAP={5,6}; CAP={5:20,6:45}
thin=Side(style='thin', color='D9DEE4'); border=Border(thin,thin,thin,thin); made=[]
for marca,ciudad in zones:
    rows=[[clean(v) for v in r] for r in q('SELECT creado_en,nombre,telefono,tipo_cliente,solicitud,detalle,asesor FROM leads WHERE marca="%s" AND ciudad="%s" ORDER BY creado_en'%(marca,ciudad))]
    accent = TEAL if marca=='Ardisa' else AMBER
    wb=Workbook(); ws=wb.active; ws.title='Reporte'; ws.sheet_view.showGridLines=False
    ws.row_dimensions[1].height=46
    try:
        lp,(lw,lh)=logo_png(marca); img=XLImage(lp); H=58; img.width=int(lw*H/lh); img.height=H; ws.add_image(img,'A1')
    except Exception: pass
    ws.cell(row=1,column=3,value='Reporte de Solicitudes de Clientes').font=Font(size=16,bold=True,color=NAVY)
    ws.cell(row=1,column=3).alignment=Alignment(vertical='center')
    ws.cell(row=2,column=3,value='%s %s   ·   Clientes que escribieron por WhatsApp   ·   Total: %d'%(marca,ciudad,len(rows))).font=Font(size=11,color=GREY)
    for i in range(1,8): ws.cell(row=3,column=i).fill=PatternFill('solid',fgColor=accent)
    ws.row_dimensions[3].height=5
    hr=4
    for i,c in enumerate(COLS,1):
        cell=ws.cell(row=hr,column=i,value=c); cell.fill=PatternFill('solid',fgColor=NAVY)
        cell.font=Font(bold=True,color='FFFFFF'); cell.alignment=Alignment(horizontal='center',vertical='center'); cell.border=border
    ws.row_dimensions[hr].height=22
    for ri,row in enumerate(rows, hr+1):
        for ci,val in enumerate(row,1):
            cell=ws.cell(row=ri,column=ci,value=val); cell.border=border
            cell.alignment=Alignment(vertical='center',wrap_text=(ci in WRAP))
            if (ri-hr)%2==0: cell.fill=PatternFill('solid',fgColor=LIGHT)
    for ci in range(1,8):
        m=len(COLS[ci-1])
        for r in rows: m=max(m, len(str(r[ci-1])))
        ws.column_dimensions[get_column_letter(ci)].width=max(min(m+2, CAP.get(ci,55)),10)
    sr=hr+len(rows)+2
    ws.cell(row=sr,column=1,value='Resumen por asesor').font=Font(bold=True,color=NAVY)
    for j,(a,n) in enumerate(Counter(r[6] for r in rows).items()):
        ws.cell(row=sr+1+j,column=1,value=a); ws.cell(row=sr+1+j,column=2,value=n)
    fn=OUT+('%s_%s.xlsx'%(marca,ciudad)).replace(' ','_'); wb.save(fn); made.append((marca,ciudad,len(rows)))
for m,c,n in made: print('  %s %s: %d'%(m,c,n))
load_workbook(OUT+'Carpincentro_Bucaramanga.xlsx'); print('round-trip OK')
