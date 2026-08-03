#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Reporte SEMANAL de leads del bot WhatsApp -> correo a los responsables (Microsoft 365 / smtp.office365.com).
# La clave NUNCA va aquí: se lee de /home/ubuntu/.config/ardisa/smtp_pass (chmod 600).
# Uso:  python3 reporte_semanal.py           -> envía a los destinatarios reales
#       python3 reporte_semanal.py --test    -> envía SOLO a la cuenta de prueba (para previsualizar)
import subprocess, os, sys, ssl, smtplib, datetime, time
from email.message import EmailMessage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === REINTENTOS SMTP (fix 2026-08-03) ===
# El 3-ago 07:00 Office365 cortó la conexión ("read operation timed out") y el script murió: Nancy, Paola y
# María NO recibieron el reporte de esa semana. Una red que falla una vez es normal; perder el correo, no.
INTENTOS, ESPERA = 3, 20        # 3 intentos, esperando 20s, 40s entre ellos (espera creciente / backoff)

def conectar():
    """Abre la sesión SMTP. Si falla, reintenta con espera creciente en vez de tumbar todo el reporte."""
    ultimo = None
    for i in range(1, INTENTOS+1):
        try:
            s = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=60)
            s.ehlo(); s.starttls(context=ssl.create_default_context()); s.ehlo()
            s.login(SMTP_USER, SMTP_PASS)
            return s
        except Exception as e:
            ultimo = e; print("SMTP conectar intento %d/%d falló: %s" % (i, INTENTOS, e))
            if i < INTENTOS: time.sleep(ESPERA*i)
    raise ultimo

def enviar(s, msg, from_addr, to_addrs):
    """Envía. Si la conexión se cayó a mitad, la reabre y vuelve a intentar. Devuelve la sesión vigente."""
    for i in range(1, INTENTOS+1):
        try:
            s.send_message(msg, from_addr=from_addr, to_addrs=to_addrs); return s
        except Exception as e:
            print("SMTP enviar intento %d/%d falló: %s" % (i, INTENTOS, e))
            if i >= INTENTOS: raise
            time.sleep(ESPERA*i)
            try: s.quit()
            except Exception: pass          # la conexión ya estaba muerta; da igual
            s = conectar()
    return s

# ---- Config ----
SMTP_HOST, SMTP_PORT = "smtp.office365.com", 587
SMTP_USER = "noreply@ardisa.com"
SMTP_PASS = open("/home/ubuntu/.config/ardisa/smtp_pass").read().strip()
# Destinatarios por MARCA (decisión Deicy 2026-07-15):
#   Ardisa (CyR)  -> Nancy (CyR) + Paola + Maria Camila (coordinacion)
#   Carpincentro  -> SOLO Paola + Maria Camila (Nancy es de Ardisa, no recibe Carpincentro)
DEST_ARDISA = ["nancy.zambrano@ardisa.com", "paola.calderon@ardisa.com", "maria.ardila@ardisa.com"]
DEST_CARP   = ["paola.calderon@ardisa.com", "maria.ardila@ardisa.com"]
MARCAS = [("Ardisa", "Ardisa", DEST_ARDISA), ("Carpincentro", "Carpincentro", DEST_CARP)]
DEST_PRUEBA = ["ernesto.rondano@ardisa.com"]   # para --test (correo de Deicy)
# Copia OCULTA (BCC) de supervisión, solo en modo real: los reales (Nancy/Paola/María) NO la ven (decisión Ernesto 2026-07-21).
BCC_COPIA = ["ernesto.rondano@ardisa.com", "deicy.jejen@ardisa.com"]
DIAS = 7
OUT = "/home/ubuntu/whatsapp-ardisa/reportes/"
os.makedirs(OUT, exist_ok=True)
NAVY, TEAL, LIGHT, GREY = "1E2A4A", "0F9D8E", "F4F6F8", "5A6472"

def q(sql):
    out = subprocess.check_output(
        ["sudo","-n","mysql","--default-character-set=utf8mb4","bot_ardisa","-N","-B","-e",sql],
        encoding="utf-8", errors="replace")
    return [ln.split("\t") for ln in out.strip().split("\n") if ln.strip()]

def build_xlsx(path, marca, titulo):
    rows = q("SELECT DATE_FORMAT(creado_en,'%%Y-%%m-%%d %%H:%%i'), nombre, telefono, ciudad, marca, "
             "tipo_cliente, solicitud, detalle, asesor FROM leads "
             "WHERE creado_en >= (NOW() - INTERVAL %d DAY) AND marca='%s' ORDER BY creado_en DESC" % (DIAS, marca))
    COLS = ["Fecha","Cliente","Teléfono","Ciudad","Marca","Perfil","Solicitud","Detalle","Asesor"]
    WRAP = {7,8}; CAP = {7:22,8:50}
    thin = Side(style="thin", color="D9DEE4"); border = Border(thin,thin,thin,thin)
    wb = Workbook(); ws = wb.active; ws.title = "Reporte semanal"; ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Grupo Ardisa · %s — Reporte semanal (WhatsApp)" % titulo).font = Font(size=15, bold=True, color=NAVY)
    ws.cell(row=2, column=1, value="Últimos %d días   ·   Total: %d clientes" % (DIAS, len(rows))).font = Font(size=11, color=GREY)
    hr = 4
    for i,c in enumerate(COLS,1):
        cell = ws.cell(row=hr, column=i, value=c); cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color="FFFFFF"); cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = border
    ws.row_dimensions[hr].height = 22
    for ri,row in enumerate(rows, hr+1):
        for ci,val in enumerate(row,1):
            cell = ws.cell(row=ri, column=ci, value=val); cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(ci in WRAP))
            if (ri-hr) % 2 == 0: cell.fill = PatternFill("solid", fgColor=LIGHT)
    for ci in range(1,len(COLS)+1):
        m = len(COLS[ci-1])
        for r in rows: m = max(m, len(str(r[ci-1])) if ci-1 < len(r) else 0)
        ws.column_dimensions[get_column_letter(ci)].width = max(min(m+2, CAP.get(ci,26)), 10)
    wb.save(path)
    return len(rows)

def main():
    test = "--test" in sys.argv
    hoy = datetime.date.today().isoformat()
    s = conectar()
    # Un reporte POR MARCA, a sus propios destinatarios (Carpincentro NO va a Nancy).
    for marca, titulo, dest_real in MARCAS:
        dest = DEST_PRUEBA if test else dest_real
        fn = OUT + "Reporte_%s_%s.xlsx" % (marca, hoy)
        n = build_xlsx(fn, marca, titulo)
        msg = EmailMessage()
        msg["From"] = "Grupo Ardisa (Bot WhatsApp) <%s>" % SMTP_USER
        msg["To"] = ", ".join(dest)
        msg["Subject"] = ("[PRUEBA] " if test else "") + "Reporte semanal %s (WhatsApp) — %s" % (titulo, hoy)
        msg.set_content(
            "Buen día,\n\nAdjuntamos el reporte semanal de solicitudes de %s recibidas por el bot de WhatsApp "
            "(últimos %d días): %d clientes.\n\nEste correo es automático, por favor no responder.\n\nGrupo Ardisa" % (titulo, DIAS, n))
        with open(fn, "rb") as f:
            msg.add_attachment(f.read(), maintype="application",
                               subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               filename=os.path.basename(fn))
        bcc = [] if test else [b for b in BCC_COPIA if b not in dest]   # BCC de supervisión (no en --test, sin duplicar a quien ya está en To)
        s = enviar(s, msg, SMTP_USER, list(dest) + bcc)   # to_addrs explícito: el BCC recibe SIN aparecer en las cabeceras
        print("OK: [%s] %d leads -> %s%s" % (titulo, n, ", ".join(dest), (" (BCC: " + ", ".join(bcc) + ")") if bcc else ""))
    s.quit()

if __name__ == "__main__":
    main()
