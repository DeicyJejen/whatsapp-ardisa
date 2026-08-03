#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Reporte SEMANAL de SEGUIMIENTO — UN solo Excel COMPLETO con las 21 columnas del formato de ellos
# (el bot llena lo que sabe; las columnas manuales de MKT quedan en blanco para completarlas en el mismo archivo).
# La clave SMTP se lee de /home/ubuntu/.config/ardisa/smtp_pass (chmod 600).
# MODO PRUEBA: por ahora solo le llega a Deicy. Al aprobar, cambiar DEST a los responsables reales.
import subprocess, os, sys, ssl, smtplib, datetime, time
from email.message import EmailMessage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from collections import Counter
from PIL import Image, ImageChops

# === RUTA DE LOS LOGOS (fix 2026-08-03) ===
# El 22-jul los archivos de marca se movieron a oficina/ y estos scripts seguían apuntando a la raíz -> desde
# entonces el correo y el Excel salieron SIN logo (el try/except lo tapaba). Ahora se buscan en varias carpetas.
LOGO_DIRS = ["/home/ubuntu/whatsapp-ardisa/oficina/", "/home/ubuntu/whatsapp-ardisa/"]
def _ruta(nombre):
    """Devuelve la PRIMERA carpeta donde el archivo exista de verdad. Si no está en ninguna, avisa fuerte."""
    for d in LOGO_DIRS:
        if os.path.exists(d + nombre): return d + nombre
    print("!! LOGO NO ENCONTRADO: %s (buscado en: %s)" % (nombre, " | ".join(LOGO_DIRS)))
    return LOGO_DIRS[-1] + nombre

# Logo por marca para el Excel (vertical, como el primer reporte). Si falta el archivo, se omite sin romper.
LOGO_MARCA = {"Ardisa": _ruta("logofirmagrupoardisavertical_org.png"),
              "Carpincentro": _ruta("logofirmacarpincentrovertical2.png")}
def _logo_sheet(marca, outdir):
    src = LOGO_MARCA.get(marca, "")
    im = Image.open(src).convert("RGBA"); bb = im.getbbox()
    if bb: im = im.crop(bb)
    bg = Image.new("RGB", im.size, "white"); bg.paste(im, mask=im.split()[3])
    o = outdir + "_logosheet_" + marca + ".png"; bg.save(o, "PNG"); return o, bg.size

LOGO_SRC  = _ruta("IMAGOTIPOS-GRUPOARDISA-01-V1-(3).jpg")            # horizontal grupoardisa (encabezado)
LOGO_VERT = _ruta("logofirmagrupoardisavertical_org.png")            # vertical (para recortar el ícono del pie)
def _trim(im):
    bg = Image.new("RGB", im.size, (255,255,255))
    bb = ImageChops.difference(im.convert("RGB"), bg).getbbox()
    return im.crop(bb) if bb else im
def _logos(outdir):
    """Logo del encabezado (imagotipo, alta resolución = nítido) + ícono del pie (la flecha del logo vertical, completa, con fondo transparente)."""
    hdr = _trim(Image.open(LOGO_SRC).convert("RGB")); hdr.thumbnail((900,900))   # alta resolución -> se ve nítido
    p_hdr = outdir+"_logo_hdr.png"; hdr.save(p_hdr,"PNG")
    v = Image.open(LOGO_VERT).convert("RGBA")   # el logo vertical trae transparencia
    bb = v.getbbox()
    if bb: v = v.crop(bb)                        # recorta el margen -> icono (arriba) + wordmark (abajo)
    W,H = v.size
    icon = v.crop((0, 0, W, int(H*0.50)))        # mitad SUPERIOR = la flecha
    bb2 = icon.getbbox()
    if bb2: icon = icon.crop(bb2)                # ajusta a la flecha
    # A PRUEBA DE RECORTE: horneamos la flecha sobre un cuadro NAVY idéntico al del pie (#1E2A4A).
    # Así el ícono es una imagen OPACA con margen -> el correo no puede recortarla y se funde con el fondo navy.
    NAVY_RGB = (30, 42, 74)
    side = int(max(icon.size) * 1.45)            # lienzo cuadrado con buen margen
    canvas = Image.new("RGB", (side, side), NAVY_RGB)
    canvas.paste(icon, ((side-icon.size[0])//2, (side-icon.size[1])//2), icon)   # flecha centrada (usa su alfa)
    canvas.thumbnail((140,140))
    p_ic = outdir+"_logo_icon.png"; canvas.save(p_ic,"PNG")
    return p_hdr, p_ic

SMTP_HOST, SMTP_PORT = "smtp.office365.com", 587
SMTP_USER = "noreply@ardisa.com"
SMTP_PASS = open("/home/ubuntu/.config/ardisa/smtp_pass").read().strip()
# EN VIVO desde 2026-08-03 (aprobado por Deicy). Mismos destinatarios y misma regla por marca que el
# reporte semanal de leads (decisión Deicy 2026-07-15): Nancy es de Ardisa, NO recibe Carpincentro.
DEST_ARDISA = ["nancy.zambrano@ardisa.com", "paola.calderon@ardisa.com", "maria.ardila@ardisa.com"]
DEST_CARP   = ["paola.calderon@ardisa.com", "maria.ardila@ardisa.com"]
DEST_PRUEBA = ["deicy.jejen@ardisa.com"]          # solo con --test
BCC_COPIA   = ["deicy.jejen@ardisa.com"]          # copia OCULTA de supervisión para Deicy (ellas no la ven)
MARCAS = [("Ardisa", DEST_ARDISA), ("Carpincentro", DEST_CARP)]   # un reporte POR línea, MISMA plantilla
DIAS = 7
OUT = "/home/ubuntu/whatsapp-ardisa/reportes/"
os.makedirs(OUT, exist_ok=True)
NAVY, TEAL, LIGHT, GREY = "1E2A4A", "0F9D8E", "F4F6F8", "5A6472"   # mismo diseño que el primer reporte (reporte.py)
AMBER = "F5B301"   # acento de Carpincentro (Ardisa usa TEAL)
C_OK, C_NO, C_WAIT, C_PEND, C_GRAY = "D8F3DF", "FBE0DE", "FCEFD2", "FDE7CC", "E7ECEC"
MESES = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

# 21 columnas EXACTAS del Excel de seguimiento de ellos
COLS = ["#","Llamada/Wapp","Mes","Fecha","Ciudad","Nombre cliente","Celular","Clasificación 01",
        "Tipo Cliente","Solicitud del cliente","Canal","Asesor","Observación Equipo comercial",
        "Valor Venta Efectiva","Estado"]
COL_ESTADO = 15   # 1-indexed
COL_VALOR  = 14
WRAP = {5,6,8,9,10,12,13,15}                # columnas de texto que hacen wrap (no se desbordan)
CAP  = {5:16, 6:20, 8:18, 9:18, 10:40, 12:22, 13:30, 14:15, 15:16}   # tope de ancho -> el texto largo baja de línea

def _fila(k,v,color):
    return ('<tr><td style="padding:9px 14px;background:#F4F7F9;border-radius:6px 0 0 6px;font-size:14px;color:#16241F">'+k+'</td>'
            '<td style="padding:9px 14px;background:#F4F7F9;border-radius:0 6px 6px 0;text-align:right;font-weight:800;color:'+color+';font-size:15px">'+v+'</td></tr>')

def q(sql):
    out = subprocess.check_output(
        ["sudo","-n","mysql","--default-character-set=utf8mb4","bot_ardisa","-N","-B","-e",sql],
        encoding="utf-8", errors="replace")
    return [ln.split("\t") for ln in out.strip().split("\n") if ln.strip()]

def estado_color(e):
    l=(e or "").lower()
    if l=="" or l=="pendiente": return C_PEND
    if "ganado" in l: return C_OK
    if "perdido" in l: return C_NO
    if "cotización" in l or "gestión" in l: return C_WAIT
    return C_GRAY

def build_xlsx(path, marca):
    raw = q("SELECT DATE_FORMAT(creado_en,'%%d/%%m/%%Y'), MONTH(creado_en), COALESCE(ciudad,''), COALESCE(nombre,''), "
            "telefono, COALESCE(solicitud,''), COALESCE(tipo_cliente,''), "
            "REPLACE(REPLACE(REPLACE(COALESCE(detalle,''),'\\n',' '),'\\r',' '),'\\t',' '), COALESCE(marca,''), "
            # Observación = "Motivo: X · <obs del asesor>" (el motivo de pérdida va aquí, NO pegado al Estado — Deicy 2026-07-21)
            "COALESCE(asesor,''), TRIM(BOTH ' · ' FROM CONCAT(IF(COALESCE(estado_motivo,'')='','',CONCAT('Motivo: ',estado_motivo,' · ')), REPLACE(REPLACE(REPLACE(COALESCE(obs_asesor,''),'\\n',' '),'\\r',' '),'\\t',' '))), "
            "COALESCE(CAST(valor_venta AS CHAR),''), COALESCE(estado,'') "
            # Últimos 7 días INCLUYENDO hoy, SOLO de esta marca (un reporte por línea).
            "FROM leads WHERE creado_en >= (CURDATE() - INTERVAL 6 DAY) AND COALESCE(modo_prueba,0)=0 "
            "AND marca='%s' ORDER BY creado_en ASC" % marca)
    raw = [ (r+[""]*13)[:13] for r in raw ]
    # fila de 21 columnas: (idx crudos) 0 fecha,1 mes,2 ciudad,3 nombre,4 tel,5 solicitud,6 tipo,7 detalle,8 marca,9 asesor,10 obs,11 valor,12 estado
    rows=[]
    for i,r in enumerate(raw,1):
        try: mes = MESES[int(r[1])]
        except: mes = ""
        estado = r[12] if r[12].strip() else "Pendiente"
        valor = ""
        if r[11].strip():
            try: valor = "$"+format(int(float(r[11])),",d").replace(",",".")
            except: valor = r[11]
        # sin emojis/pictogramas: en Excel de escritorio salen en blanco y negro o como cuadros (Deicy 2026-07-21)
        import re as _re
        _clean = lambda v: _re.sub(r' {2,}',' ', _re.sub(r'[\U0001F000-\U0001FAFF☀-➿⬀-⯿←-⇿️‍⃣ℹ]', ' ', v)).strip() if isinstance(v,str) else v
        rows.append([ _clean(x) for x in [ i, "WhatsApp", mes, r[0], r[2], r[3], r[4], r[5], r[6], r[7],
                      r[8], r[9], r[10], valor, estado ] ])
    # métricas (de los crudos)
    tot=len(raw); rep=sum(1 for r in raw if r[12].strip()); pend=tot-rep
    gan=sum(1 for r in raw if "ganado" in r[12].lower())
    val=0.0
    for r in raw:
        if "ganado" in r[12].lower():
            try: val+=float(r[11] or 0)
            except: pass
    thin=Side(style="thin",color="D9DEE4"); border=Border(thin,thin,thin,thin)
    accent = TEAL if marca=="Ardisa" else AMBER   # acento por marca (Ardisa teal, Carpincentro ámbar)
    wb=Workbook(); ws=wb.active; ws.title="Seguimiento"; ws.sheet_view.showGridLines=False
    # --- encabezado: logo de la marca (A1) + título + subtítulo, igual que el primer reporte ---
    ws.row_dimensions[1].height=46
    try:
        lp,(lw,lh)=_logo_sheet(marca, OUT); img=XLImage(lp); _H=58; img.width=int(lw*_H/lh); img.height=_H; ws.add_image(img,"A1")
    except Exception as _e:
        print("aviso logo Excel (%s): %s" % (marca,_e))
    ws.cell(row=1,column=3,value="Reporte de Solicitudes de Clientes").font=Font(size=16,bold=True,color=NAVY)
    ws.cell(row=1,column=3).alignment=Alignment(vertical="center")
    ws.cell(row=2,column=3,value="%s  ·  Seguimiento de solicitudes por WhatsApp  ·  Últimos 7 días  ·  Total: %d" % (marca,tot)).font=Font(size=11,color=GREY)
    for i in range(1,len(COLS)+1): ws.cell(row=3,column=i).fill=PatternFill("solid",fgColor=accent)   # línea de acento (sutil)
    ws.row_dimensions[3].height=3
    hr=4
    for i,c in enumerate(COLS,1):
        cell=ws.cell(row=hr,column=i,value=c); cell.fill=PatternFill("solid",fgColor=NAVY)
        cell.font=Font(bold=True,color="FFFFFF"); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border
    ws.row_dimensions[hr].height=22
    for ri,row in enumerate(rows,hr+1):
        for ci,v in enumerate(row,1):
            cell=ws.cell(row=ri,column=ci,value=v); cell.border=border
            cell.alignment=Alignment(vertical="center",wrap_text=(ci in WRAP),horizontal=("center" if ci in (1,2,3) else "left"))
            if ci==COL_ESTADO:
                cell.fill=PatternFill("solid",fgColor=estado_color(row[COL_ESTADO-1])); cell.font=Font(bold=True)
            elif (ri-hr)%2==0:
                cell.fill=PatternFill("solid",fgColor=LIGHT)
    widths={}
    for ci in range(1,len(COLS)+1):
        m=len(COLS[ci-1])
        for r in rows: m=max(m,len(str(r[ci-1])) if ci-1<len(r) else 0)
        w=max(min(m+2,CAP.get(ci,26)),9); widths[ci]=w
        ws.column_dimensions[get_column_letter(ci)].width=w
    # ALTURA DE FILA automática: según el texto que hace wrap, para que NADA se corte ni se desborde.
    for ri,row in enumerate(rows,hr+1):
        lineas=1
        for ci in WRAP:
            _txt=str(row[ci-1]) if ci-1<len(row) else ""
            cw=max(widths.get(ci,20)-1, 8)
            lineas=max(lineas, -(-len(_txt)//cw))   # ceil(len/ancho) = nº de renglones que ocupa
        ws.row_dimensions[ri].height=min(max(17, lineas*15.0), 150)
    # --- Resumen por asesor (como el primer reporte) ---
    sr=hr+len(rows)+2
    ws.cell(row=sr,column=1,value="Resumen por asesor").font=Font(bold=True,color=NAVY)
    for j,(a,n) in enumerate(Counter(r[11] for r in rows if r[11]).items()):
        ws.cell(row=sr+1+j,column=1,value=a); ws.cell(row=sr+1+j,column=2,value=n)
    ws.freeze_panes="A%d"%(hr+1)
    wb.save(path)
    return tot,rep,pend,gan,int(val)

def _pl(n, sing, plur):
    return "<b>%d</b> %s" % (n, sing if n==1 else plur)

def _resumen(tot, rep, pend, gan):
    return ("Resumen del período: " + _pl(tot,"solicitud recibida","solicitudes recibidas")
            + " &middot; " + _pl(rep,"con resultado reportado","con resultado reportado")
            + " &middot; " + _pl(pend,"pendiente por reportar","pendientes por reportar")
            + " &middot; " + _pl(gan,"venta ganada","ventas ganadas") + ".")

def _html_correo(marca, tot, rep, pend, gan):
    return ('<div style="margin:0;padding:24px 12px;background:#EEF1F4;font-family:\'Segoe UI\',Arial,sans-serif">'
      '<table role="presentation" cellpadding="0" cellspacing="0" align="center" width="620" style="width:620px;max-width:100%;margin:0 auto;background:#ffffff;border-radius:10px;overflow:hidden;box-shadow:0 1px 4px rgba(20,40,80,.08)">'
      # --- encabezado: título + logo grupoardisa ---
      '<tr><td style="padding:26px 30px 18px">'
        '<table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>'
          '<td style="vertical-align:middle">'
            '<div style="font-size:19px;font-weight:600;color:#1E2A4A;line-height:1.25">Reporte de Solicitudes de Clientes</div>'
            '<div style="font-size:12px;font-weight:400;color:#9AA6B2;margin-top:5px">Seguimiento semanal &middot; ' + marca + '</div>'
          '</td>'
          '<td align="right" style="vertical-align:middle"><img src="cid:logohdr" width="146" style="display:block;width:146px;height:auto;border:0"></td>'
        '</tr></table>'
      '</td></tr>'
      # --- línea teal ---
      '<tr><td style="padding:0 30px"><div style="height:2px;background:#0F9D8E;border-radius:2px;font-size:0;line-height:0">&nbsp;</div></td></tr>'
      # --- cuerpo ---
      '<tr><td style="padding:32px 30px 8px;color:#2A3340;font-size:14px;line-height:1.6">'
        '<p style="margin:0 0 12px">Cordial saludo,</p>'
        '<p style="margin:0 0 12px">Ponemos a su disposición el informe de seguimiento de las <b>solicitudes de clientes de ' + marca + '</b> recibidas a través de nuestro canal de WhatsApp durante los <b>últimos 7 días</b>.</p>'
        '<p style="margin:0 0 14px">El informe relaciona, para cada solicitud, la fecha, el cliente, la ciudad, la clasificación, el asesor asignado y el estado de gestión, junto con el valor de la venta y las observaciones registradas por el equipo comercial.</p>'
        '<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="margin:6px 0 14px"><tr>'
          '<td style="border-left:4px solid #0F9D8E;background:#F4F6F8;padding:12px 16px;color:#2A3340;font-size:14px;border-radius:0 6px 6px 0">&#128206;&nbsp; El detalle completo se encuentra en el archivo de <b>Excel adjunto</b>.</td>'
        '</tr></table>'
        '<p style="margin:0">Cordialmente,</p>'
        '<p style="margin:2px 0 0;font-weight:700;color:#1E2A4A;font-size:14px">Grupo Ardisa</p>'
      '</td></tr>'
      # --- pie navy con dirección ---
      '<tr><td style="padding:16px 28px 20px">'
        '<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#1E2A4A;border-radius:10px"><tr>'
          '<td width="66" style="padding:12px 0 12px 18px;vertical-align:middle"><img src="cid:logoicon" width="44" height="44" style="display:block;width:44px;height:44px;border:0"></td>'
          '<td style="padding:15px 18px 15px 6px;vertical-align:middle">'
            '<div style="color:#2FBFA8;font-size:15px;font-weight:800">Grupo Ardisa</div>'
            '<div style="color:#D6DEEA;font-size:12px;margin-top:2px">Cra. 17C No. 60 - 30 Bucaramanga &middot; Cel. (+57) 3188640235 - 3144714359</div>'
            '<div style="margin-top:2px"><a href="https://www.ardisa.com" style="color:#4FD1C5;font-size:12px;text-decoration:none">www.ardisa.com</a></div>'
            '<div style="color:#8894A3;font-size:11px;margin-top:6px">Grupo Ardisa &middot; Correo automático &middot; No responder a este mensaje</div>'
          '</td>'
        '</tr></table>'
      '</td></tr>'
      '</table></div>')

def enviar_marca(s, marca, dest, test, hoy):
    fn = OUT + "Seguimiento_%s_%s.xlsx" % (marca, hoy)
    tot,rep,pend,gan,val = build_xlsx(fn, marca)
    msg=EmailMessage()
    msg["From"]="Grupo Ardisa (Bot WhatsApp) <%s>"%SMTP_USER
    msg["To"]=", ".join(dest)
    msg["Subject"]=("[PRUEBA] " if test else "")+"Seguimiento semanal de solicitudes — %s (WhatsApp) — %s"%(marca,hoy)
    msg.set_content(
        "Cordial saludo,\n\n"
        "Ponemos a su disposición el informe de seguimiento de las solicitudes de clientes de %s "
        "recibidas a través de nuestro canal de WhatsApp durante los últimos 7 días.\n\n"
        "El informe relaciona, para cada solicitud, la fecha, el cliente, la ciudad, la clasificación, "
        "el asesor asignado y el estado de gestión, junto con el valor de la venta y las observaciones "
        "registradas por el equipo comercial. El detalle completo se encuentra en el archivo de Excel adjunto.\n\n"
        "Cordialmente,\nGrupo Ardisa\n\n"
        "Este es un correo automático, por favor no responder."
        % marca)
    msg.add_alternative(_html_correo(marca, tot, rep, pend, gan), subtype='html')
    _hpart = msg.get_payload()[-1]   # imágenes en línea (CID): logo encabezado + ícono pie
    try:
        _ph, _pi = _logos(OUT)
        # cid CON signos <>: el estándar (RFC 2045) exige "Content-ID: <logohdr>"; Python NO los pone solo
        # y Outlook puede no emparejarlo con el src="cid:logohdr" del HTML -> imagen rota. (fix 2026-08-03)
        with open(_ph,"rb") as _f: _hpart.add_related(_f.read(), "image", "png", cid="<logohdr>")
        with open(_pi,"rb") as _f: _hpart.add_related(_f.read(), "image", "png", cid="<logoicon>")
    except Exception as _e:
        print("aviso logos:", _e)
    with open(fn,"rb") as f:
        msg.add_attachment(f.read(),maintype="application",
                           subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           filename=os.path.basename(fn))
    bcc = [] if test else [b for b in BCC_COPIA if b not in dest]   # copia oculta de supervisión (no en --test)
    for i in range(1, INTENTOS+1):          # si la conexión se cayó a mitad, se reabre y se reintenta
        try:
            s.send_message(msg, from_addr=SMTP_USER, to_addrs=list(dest)+bcc); break
        except Exception as e:
            print("SMTP enviar intento %d/%d falló: %s" % (i, INTENTOS, e))
            if i >= INTENTOS: raise
            time.sleep(ESPERA*i)
            try: s.quit()
            except Exception: pass
            s = conectar()
    print("OK: %s — %d solicitudes (%d reportadas, %d pend) -> %s%s"%(marca,tot,rep,pend,", ".join(dest),
          (" (BCC: "+", ".join(bcc)+")") if bcc else ""))

INTENTOS, ESPERA = 3, 20        # fix 2026-08-03: Office365 corta la conexión de vez en cuando -> reintentar
def conectar():
    """Abre la sesión SMTP con reintentos y espera creciente (backoff) para no perder el reporte."""
    ultimo = None
    for i in range(1, INTENTOS+1):
        try:
            s = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=60)
            s.ehlo(); s.starttls(context=ssl.create_default_context()); s.ehlo()
            s.login(SMTP_USER, SMTP_PASS)
            return s
        except Exception as e:
            ultimo = e; print("SMTP conectar intento %d/%d falló: %s" % (i, INTENTOS, e))
            if i < INTENTOS: time.sleep(ESPERA*i)
    raise ultimo

def main():
    test="--test" in sys.argv
    hoy=datetime.date.today().isoformat()
    s=conectar()
    for marca, dest_real in MARCAS:            # un reporte POR línea, MISMA plantilla
        dest = DEST_PRUEBA if test else dest_real
        enviar_marca(s, marca, dest, test, hoy)
    s.quit()

if __name__=="__main__":
    main()
